"""
Génère le dossier d'investissement (Excel) : synthèse, flux de trésorerie,
grille de sensibilité NPV (2 variables), hypothèses.

Usage : python src/report.py
"""
from dataclasses import replace
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from model import Assumptions, evaluate, fetch_baseline, npv, cash_flows, self_check

OUT = Path(__file__).resolve().parent.parent / "output"

PETROL = "137A8B"; PETROL_D = "0C5563"
GREEN = "2FA36B"; RED = "D9534F"
EURO = '#,##0" €"'
thin = Side(style="thin", color="DCE2E8")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HEAD_FILL = PatternFill("solid", fgColor=PETROL)
HEAD_FONT = Font(bold=True, color="FFFFFF")

REDUCTION_GRID = [0.005, 0.01, 0.015, 0.02, 0.025, 0.03, 0.035]
MARGIN_GRID = [0.15, 0.20, 0.25, 0.30, 0.35]


def header_row(ws, row, headers, start=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start + i, value=h)
        c.fill = HEAD_FILL; c.font = HEAD_FONT; c.border = BORDER
        c.alignment = Alignment(horizontal="center")


def sensitivity_grid(base, a):
    grid = {}
    for rp in REDUCTION_GRID:
        for gm in MARGIN_GRID:
            flows = cash_flows(base, replace(a, reduction_pts=rp, gross_margin_pct=gm))
            grid[(rp, gm)] = npv(flows, a.discount_rate)
    return grid


def build(base, a, result, path):
    wb = Workbook()

    ws = wb.active; ws.title = "Synthèse"; ws.sheet_view.showGridLines = False
    ws["B2"] = "Business case — dispositif anti-annulation de commande"
    ws["B2"].font = Font(bold=True, size=16, color=PETROL_D)
    ws["B3"] = f"Généré automatiquement le {date.today().isoformat()}"
    ws["B3"].font = Font(italic=True, color="5E6B7A")

    ws["B5"] = "Constat mesuré (base réelle, Projet 07)"
    ws["B5"].font = Font(bold=True, color="5E6B7A", size=10)
    ws["B6"] = (f"{base.annual_orders:,.0f} commandes/an · {base.cancellation_rate:.1%} annulées "
                f"· valeur moyenne d'une commande annulée {base.avg_value_cancelled:,.0f} €")

    cards = [("NPV (VAN)", result.npv, EURO),
             ("IRR (TRI)", result.irr, '0.0%'),
             ("Payback", result.payback_years, '0.0" ans"')]
    for i, (label, val, fmt) in enumerate(cards):
        col = 2 + i * 2
        lc = ws.cell(row=8, column=col, value=label)
        lc.font = Font(bold=True, color="5E6B7A", size=9)
        vc = ws.cell(row=9, column=col, value=val)
        vc.font = Font(bold=True, size=18, color=(GREEN if (val or 0) > 0 else RED))
        vc.number_format = fmt
        ws.merge_cells(start_row=8, start_column=col, end_row=8, end_column=col + 1)
        ws.merge_cells(start_row=9, start_column=col, end_row=9, end_column=col + 1)
    for col in range(2, 10):
        ws.column_dimensions[get_column_letter(col)].width = 16

    ws_f = wb.create_sheet("Flux de trésorerie"); ws_f.sheet_view.showGridLines = False
    header_row(ws_f, 1, ["Année", "Flux net €", "Flux cumulé €"])
    cumulative = 0
    for year, cf in enumerate(result.cash_flows):
        cumulative += cf
        ws_f.cell(row=year + 2, column=1, value=year).border = BORDER
        c1 = ws_f.cell(row=year + 2, column=2, value=round(cf)); c1.number_format = EURO; c1.border = BORDER
        c2 = ws_f.cell(row=year + 2, column=3, value=round(cumulative)); c2.number_format = EURO; c2.border = BORDER
    ws_f.column_dimensions["A"].width = 10; ws_f.column_dimensions["B"].width = 16; ws_f.column_dimensions["C"].width = 16
    bar = BarChart(); bar.title = "Flux net par année"; bar.height = 8; bar.width = 16; bar.legend = None
    data = Reference(ws_f, min_col=2, min_row=1, max_row=len(result.cash_flows) + 1)
    cats = Reference(ws_f, min_col=1, min_row=2, max_row=len(result.cash_flows) + 1)
    bar.add_data(data, titles_from_data=True); bar.set_categories(cats)
    bar.series[0].graphicalProperties.solidFill = PETROL
    ws_f.add_chart(bar, "E2")

    ws_s = wb.create_sheet("Sensibilité NPV"); ws_s.sheet_view.showGridLines = False
    ws_s["B2"] = "NPV (€) selon le taux de réduction des annulations (lignes) et la marge brute (colonnes)"
    ws_s["B2"].font = Font(bold=True, size=11, color=PETROL_D)
    grid = sensitivity_grid(base, a)
    header_row(ws_s, 4, [""] + [f"{gm:.0%}" for gm in MARGIN_GRID], start=2)
    for r, rp in enumerate(REDUCTION_GRID, start=5):
        ws_s.cell(row=r, column=2, value=f"{rp:.1%}").font = Font(bold=True); ws_s.cell(row=r, column=2).border = BORDER
        for c, gm in enumerate(MARGIN_GRID, start=3):
            cell = ws_s.cell(row=r, column=c, value=round(grid[(rp, gm)]))
            cell.number_format = EURO; cell.border = BORDER
    ws_s.column_dimensions["B"].width = 12
    for c in range(3, 3 + len(MARGIN_GRID)):
        ws_s.column_dimensions[get_column_letter(c)].width = 13
    last_row = 4 + len(REDUCTION_GRID)
    last_col = get_column_letter(2 + len(MARGIN_GRID))
    ws_s.conditional_formatting.add(
        f"C5:{last_col}{last_row}",
        ColorScaleRule(start_type="min", start_color=RED, mid_type="num", mid_value=0, mid_color="FFFFFF",
                        end_type="max", end_color=GREEN))

    ws_h = wb.create_sheet("Hypothèses"); ws_h.sheet_view.showGridLines = False
    header_row(ws_h, 1, ["Hypothèse", "Valeur", "Origine"])
    rows = [
        ("Commandes / an", f"{base.annual_orders:,.0f}", "Mesuré (Projet 07, 24 mois complets)"),
        ("Taux d'annulation actuel", f"{base.cancellation_rate:.2%}", "Mesuré (Projet 07)"),
        ("Valeur moy. commande annulée", f"{base.avg_value_cancelled:,.0f} €", "Mesuré (Projet 07)"),
        ("Points d'annulation récupérés", f"{a.reduction_pts:.1%}", "Hypothèse (dossier fournisseur / benchmark secteur)"),
        ("Marge brute", f"{a.gross_margin_pct:.0%}", "Hypothèse (non présente dans le schéma OLTP)"),
        ("Investissement initial", f"{a.investment_cost:,.0f} €", "Hypothèse (implémentation + intégration + formation)"),
        ("Coût annuel du dispositif", f"{a.annual_program_cost:,.0f} €", "Hypothèse (licence + ~0,3 ETP exploitation)"),
        ("Taux d'actualisation", f"{a.discount_rate:.0%}", "Hypothèse (coût du capital)"),
        ("Horizon", f"{a.horizon_years} ans", "Choix de modélisation"),
        ("Montée en charge", str(a.ramp_up), "Hypothèse (déploiement progressif)"),
    ]
    for r, (label, val, origin) in enumerate(rows, start=2):
        ws_h.cell(row=r, column=1, value=label).border = BORDER
        ws_h.cell(row=r, column=2, value=val).border = BORDER
        ws_h.cell(row=r, column=3, value=origin).border = BORDER
    ws_h.column_dimensions["A"].width = 32; ws_h.column_dimensions["B"].width = 16; ws_h.column_dimensions["C"].width = 48

    wb.save(path)


def main() -> None:
    OUT.mkdir(exist_ok=True)
    base = fetch_baseline()
    a = Assumptions()
    self_check(base, a)
    result = evaluate(base, a)
    path = OUT / f"business_case_{date.today().isoformat()}.xlsx"
    build(base, a, result, path)
    print(f"Rapport généré : {path.name}")
    print(f"  NPV {result.npv:,.0f} € · IRR {result.irr:.1%} · payback {result.payback_years:.1f} ans")


if __name__ == "__main__":
    main()
